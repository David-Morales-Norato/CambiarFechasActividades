# -*- coding: utf-8 -*-
"""
Spyder Editor

This is a temporary script file.

By David A. Miranda, PhD (2017)




'weekDay' are:
    0 to Monday
    1 to Tuesday
    2 to Wednesday
    3 to Thursday
    4 to Friday
    5 to Saturday
    6 to Sunday

'start' is in 24 hour format, i.e. a number from 0 to 23
"""
from openpyxl import load_workbook

def getCPSDataFromXLS(coursesCPSActivationData):
    
    logs = ''
    
    cpsXLSData = load_workbook(coursesCPSActivationData)
    
    swError = False
    courses = []
    cps = []
    calendar = {}
    
    if 'courses' in cpsXLSData.sheetnames:
        sCourses = cpsXLSData['courses']
        NRows = sCourses.max_row
        NColumns = sCourses.max_column
        for row in range(2,NRows+1):
            courses.append({
                    'group':   sCourses.cell(row=row, column=1).value,
                    'teacher': sCourses.cell(row=row, column=2).value,
                    'link':    sCourses.cell(row=row, column=3).value,
                    'weekDay': sCourses.cell(row=row, column=4).value,
                    'hour':    sCourses.cell(row=row, column=5).value
                    })
    else:
        swError = True
        logs += '\nError: courses sheet do not exist in ' + coursesCPSActivationData
    
    if 'cps' in cpsXLSData.sheetnames:
        sCourses = cpsXLSData['cps']
        NRows = sCourses.max_row
        NColumns = sCourses.max_column
        for row in range(2,NRows+1):
            cps.append({
                    'courseName': sCourses.cell(row=row, column=1).value,
                    'cpsName':    sCourses.cell(row=row, column=2).value,
                    'module':     sCourses.cell(row=row, column=3).value,
                    'week':       sCourses.cell(row=row, column=4).value,
                    'attempts':   sCourses.cell(row=row, column=5).value
                    })
    else:
        swError = True
        logs += '\nError: cps sheet do not exist in ' + coursesCPSActivationData
        
    if 'calendar' in cpsXLSData.sheetnames:
        sCourses = cpsXLSData['calendar']
        NRows = sCourses.max_row
        NColumns = sCourses.max_column
        for row in range(2,NRows+1):
            calendar[sCourses.cell(row=row, column=1).value] = sCourses.cell(row=row, column=2).value
    else:
        swError = True
        logs += '\nError: calendar sheet do not exist in ' + coursesCPSActivationData
    
    return courses, cps, calendar, swError, cpsXLSData, logs
